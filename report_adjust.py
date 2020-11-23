import os
import openpyxl
import tkinter
from tkinter import messagebox


def run():
    if os.path.exists("settings.txt"):
        setting = open("settings.txt")
        global report_name
        global tool_name
        tool_name = setting.readline().strip()
        tool_name = tool_name.replace("tool_name=","")
        tool_name = tool_name.replace(" ","")
        report_name = setting.readline().strip()
        report_name = report_name.replace("report_name=","")
        report_name = report_name.replace(" ","")
    
    else:
        tkinter.Tk().withdraw()
        messagebox.showerror("File not found","settings.txt not found")

    if os.path.exists(report_name):
        excel_adjust()  # run method
    else:
        tkinter.Tk().withdraw()
        messagebox.showerror("File not found",report_name+" not found")


def excel_adjust():
    wb = openpyxl.load_workbook(report_name)
    sheet = wb.active
    sheet.insert_cols(1)
    current_sheet = wb.worksheets[0]
    row_count = current_sheet.max_row
    for i in range(0,int(row_count/2)):
        current_sheet.cell(row=2*i+1,column=1).value = i+1
    for i in range(0,row_count,2):  #merge row
        current_sheet.merge_cells(start_row=i+1,start_column=1,end_row=i+2,end_column=1)
        
    
    for i in range(1,sheet.max_column+1):  # column width adjust
        max_val =0
        for cell in sheet[openpyxl.utils.get_column_letter(i)]:
            val_to_check = len(str(cell.value))
            if val_to_check>max_val:
                max_val=val_to_check
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max_val+1
    for i in range(0,row_count):
        for k in range(0,current_sheet.max_column):
            if (i%4==0 )or (i%4 ==1):
                current_sheet.cell(row=i+1,column=k+1).fill = openpyxl.styles.PatternFill(start_color="FFECC9",end_color="FFECC9",fill_type="solid")
            else :
                current_sheet.cell(row=i+1,column=k+1).fill = openpyxl.styles.PatternFill(start_color="FFDEDE",end_color="FFDEDE",fill_type="solid")
            # make thick border
            current_sheet.cell(row=i+1,column=k+1).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style="thick"),right=openpyxl.styles.Side(style="thick"),top=openpyxl.styles.Side(style="thick"),bottom=openpyxl.styles.Side(style="thick"))

    wb.save(report_name)


run()